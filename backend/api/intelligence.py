from fastapi import APIRouter, Depends, BackgroundTasks, HTTPException, UploadFile, File, Form, Request, Query
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from db.models import get_db
from api.auth import get_optional_user, log_audit, require_role, institutional_access
from api.source_center import authorize_source_monitor
from db.models import User
from db.models_intelligence import (
    NationalCrimeStats,
    NationalReferenceCoverage,
    IngestionLog,
    TerritorialContext,
    IngestionFile,
    ReportRun,
    RNMCMeasure,
)
from db.models_alerts import IntelligenceAlert, IntelligenceAlertSnapshot
from services.scraper_mindefensa import MinDefensaScraper
from services.excel_processor import NationalStatsProcessor
from services.pdf_report_service import PdfReportService
from services.distribution_service import DistributionService
from services.rnmc_service import RNMCService
from services.ingest_rnmc import RNMCIngestor
from services.intelligence_service import IntelligenceService
from services.report_automation_service import ReportAutomationService
from services.alerts_rnmc import generate_rnmc_alerts
from services.alerts_prioritizer import compute_action_score, get_scoring_config
from services.ai_prioritizer import build_ai_rationale
from services.national_context_service import (
    comparable_reference_rate,
    comparable_national_rate,
    municipality_code_for_name,
    municipality_name_for_code,
    named_territorial_comparison,
    national_benchmark_guard,
    normalize_municipality_code,
    population_peer_codes,
    year_over_year,
)
from services import dq_service
from db import crud_dq
from core.config import is_strong_secret
try:
    from weasyprint import HTML, CSS
except Exception as e:
    print(f"[AVISO] WeasyPrint no disponible en inteligencia: {e}")
import logging
import hashlib
import hmac
import json
import os
from datetime import date, datetime
from io import BytesIO
from api.ia import call_gemini, call_mistral, AI_PROVIDER, GEMINI_API_KEY, MISTRAL_API_KEY
from sqlalchemy import text, func, desc
from uuid import UUID

router = APIRouter(tags=["Intelligence"])
logger = logging.getLogger("sisc_api")
REPORT_TRIGGER_ROLES = {"TI_ADMIN", "FUNC_ADMIN", "ANALYST"}
COMPACT_REFERENCE_SOURCE = "MINDEFENSA_REFERENCE_COMPACT"


class ReferenceAggregateRecord(BaseModel):
    codigo_dane: str
    municipio: str
    departamento: str
    anio: int
    mes: int
    cantidad: int


class ReferenceCoveragePeriod(BaseModel):
    anio: int
    municipality_codes: List[str]


class ReferenceAggregateUpload(BaseModel):
    filename: str
    tipo_delito: str
    records: List[ReferenceAggregateRecord]
    coverage: List[ReferenceCoveragePeriod]
    source_cutoff: Optional[date] = None


def _authorize_report_trigger(request: Request, user: Optional[User]) -> str:
    role_codes = {role.code for role in (user.roles or [])} if user else set()
    if role_codes.intersection(REPORT_TRIGGER_ROLES):
        return "USER"

    expected_key = os.getenv("SISC_REPORT_TRIGGER_KEY", "").strip()
    if not is_strong_secret(expected_key):
        raise HTTPException(
            status_code=503,
            detail="La automatizacion de reportes no esta configurada.",
        )

    provided_key = request.headers.get("X-API-KEY", "")
    if not provided_key or not hmac.compare_digest(provided_key, expected_key):
        raise HTTPException(status_code=403, detail="Acceso denegado.")
    return "SERVICE"

@router.get("/executive-brief")
async def get_executive_brief(
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    """
    Retorna un resumen ejecutivo ágil con IA y fechas de corte para Jamundí.
    """
    from sqlalchemy import func
    from db.models_intelligence import NationalCrimeStats
    from services.intelligence_service import IntelligenceService
    
    count_jamundi = db.query(func.count(NationalCrimeStats.id)).filter(
        NationalCrimeStats.municipio_normalizado.ilike('%JAMUNDI%')
    ).scalar()
    
    delitos_disponibles = db.query(NationalCrimeStats.tipo_delito).filter(
        NationalCrimeStats.municipio_normalizado.ilike('%JAMUNDI%')
    ).distinct().all()
    
    briefs = await IntelligenceService.get_executive_brief(db)
    
    return {
        "debug": {
            "total_jamundi": count_jamundi,
            "delitos": [d[0] for d in delitos_disponibles]
        },
        "briefs": briefs
    }


@router.post("/upload")
async def upload_intelligence_file(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["SOURCE_UPLOADER", "TI_ADMIN"]))
):
    """
    Carga manual de archivos Excel de MinDefensa.
    Procesa el archivo y carga los datos en la base de datos.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos Excel (.xlsx, .xls)")

    # Crear log de inicio
    log_entry = IngestionLog(
        estado="IN_PROGRESS",
        registros_insertados=0,
        errores=None,
        detalles={"filename": file.filename}
    )
    db.add(log_entry)
    db.commit()
    db.refresh(log_entry)

    try:
        import uuid
        ingestion_id = uuid.uuid4()
        contents = await file.read()
        file_hash = hashlib.sha256(contents).hexdigest()
        
        processor = NationalStatsProcessor()
        # Consumir generator para validar y obtener source_id
        all_records = list(processor.process_excel(contents, file.filename))

        detected_source_id = all_records[0].get("source_id", "GENERIC_CRIME") if all_records else "FUENTE_NO_DETECTADA"
        quality_report = dq_service.run_records_dq(
            all_records,
            file.filename or "archivo_sin_nombre",
            source_name=f"INTELLIGENCE_{detected_source_id}",
        )
        db_quality_report = crud_dq.create_dq_report(db, quality_report)
        if quality_report.get("semaforo") == "ROJO":
            log_entry.estado = "FAILED"
            log_entry.errores = "Carga bloqueada por control automatico de calidad."
            log_entry.fecha_fin = datetime.utcnow()
            log_entry.detalles = {
                "filename": file.filename,
                "dq_report_id": str(db_quality_report.id),
                "semaforo": "ROJO",
            }
            db.commit()
            raise HTTPException(
                status_code=422,
                detail={
                    "message": "El archivo fue bloqueado por errores criticos de calidad.",
                    "report_id": str(db_quality_report.id),
                    "semaforo": "ROJO",
                    "issues_count": len(quality_report.get("issues", [])),
                },
            )
        
        if not all_records:
            return {
                "ingestion_id": str(ingestion_id),
                "status": "REJECTED", 
                "message": "Archivo vacío o sin registros válidos", 
                "records": 0
            }
            
        source_id = all_records[0].get("source_id", "GENERIC_CRIME")
        
        # 0. Verificar Idempotencia
        existing_file = db.query(IngestionFile).filter(
            IngestionFile.file_hash == file_hash,
            IngestionFile.source_type == source_id
        ).first()
        
        if existing_file:
            logger.info(f"El archivo {file.filename} ya fue procesado como {source_id}.")
            return {
                "ingestion_id": str(existing_file.ingestion_id),
                "source_id": source_id,
                "file_hash": file_hash,
                "status": "REJECTED",
                "message": "Archivo ya procesado para esta fuente",
                "inserted_count": 0,
                "updated_count": 0,
                "skipped_count": existing_file.records_count,
                "periodo": existing_file.periodo_detectado,
                "quality": {
                    "report_id": str(db_quality_report.id),
                    "semaforo": quality_report.get("semaforo"),
                    "score": quality_report.get("score_overall"),
                },
            }

        # 1. Detectar Periodo y Distribución
        fechas = [r.get("fecha_hecho") for r in all_records if r.get("fecha_hecho")]
        anos = sorted(list(set([r.get("anio") for r in all_records if r.get("anio")])))
        semanas = sorted(list(set([r.get("semana") for r in all_records if r.get("semana")])))
        distribucion_anio = {}
        for r in all_records:
            a = r.get("anio")
            distribucion_anio[a] = distribucion_anio.get(a, 0) + 1
            
        periodo_min = min(fechas) if fechas else None
        periodo_max = max(fechas) if fechas else None
        periodo_str = f"{periodo_min} a {periodo_max}" if periodo_min else "Desconocido"

        # Alerta de años extraños
        alertas = []
        for a in anos:
            if a < 2000 or a > datetime.now().year + 1:
                alertas.append(f"Atención: Año detectado fuera de rango normal: {a}")

        count = 0
        res = {}
        rnmc_processed = False
        from sqlalchemy.dialects.postgresql import insert
        
        # 2. Transacción Atómica
        with db.begin_nested():
            for record_dict in all_records:
                # Forzar source_id SEM_POLICIA si es una carga SEM
                if source_id == "SEM_POLICIA":
                    record_dict["source_id"] = "SEM_POLICIA"
                    
                if record_dict.get("fuente_type") == "TERRITORIAL_CONTEXT":
                    record_dict.pop("fuente_type", None)
                    stmt = insert(TerritorialContext).values(record_dict)
                    stmt = stmt.on_conflict_do_update(
                        index_elements=['source_id', 'event_fingerprint'],
                        set_={
                            "cantidad": TerritorialContext.cantidad + record_dict["cantidad"],
                            "fuente_archivo": record_dict["fuente_archivo"]
                        }
                    )
                    db.execute(stmt)
                    count += 1
                elif source_id == "INSPECCION_MEDIDAS_RNMC":
                    if rnmc_processed:
                        continue
                    rnmc_processed = True
                    # RNMC usa su propio ingestor para lógica de fingerprints específica
                    ingestor = RNMCIngestor(db)
                    res = ingestor.process_file(contents, file.filename)
                    # Resumen de filas afectadas (insertadas + actualizadas)
                    inserted = int(res.get("inserted", 0))
                    updated = int(res.get("updated", 0))
                    total = int(res.get("total", 0))
                    count = inserted + updated
                    break
                else:
                    stmt = insert(NationalCrimeStats).values(record_dict)
                    stmt = stmt.on_conflict_do_update(
                        index_elements=['source_id', 'event_fingerprint'],
                        set_={
                            "cantidad": NationalCrimeStats.cantidad + record_dict["cantidad"],
                            "fuente_archivo": record_dict["fuente_archivo"]
                        }
                    )
                    db.execute(stmt)
                    count += 1

            # 3. Registrar archivo como procesado
            new_file_ref = IngestionFile(
                ingestion_id=ingestion_id,
                filename=file.filename,
                source_type=source_id,
                file_hash=file_hash,
                inserted_count=count,
                updated_count=0,
                skipped_count=0,
                records_count=count,
                periodo_detectado=periodo_str,
                periodo_detectado_min=periodo_min,
                periodo_detectado_max=periodo_max,
                anios_incluidos=anos,
                semanas_incluidas=semanas,
                status="COMPLETED"
            )
            db.add(new_file_ref)
            
            log_entry.estado = "COMPLETED"
            log_entry.registros_insertados = count
            log_entry.detalles = {
                "ingestion_id": str(ingestion_id),
                "source_id": source_id, 
                "filename": file.filename,
                "periodo": periodo_str,
                "distribucion": distribucion_anio,
                "alertas": alertas,
                "dq_report_id": str(db_quality_report.id),
                "dq_semaforo": quality_report.get("semaforo"),
            }

        db.commit()
        base_response = {
            "ingestion_id": str(ingestion_id),
            "source_id": source_id,
            "file_hash": file_hash,
            "status": "COMPLETED",
            "message": "Carga multi-año exitosa con trazabilidad institucional",
            "inserted_count": count,
            "periodo_detectado": periodo_str,
            "anios_incluidos": anos,
            "distribucion_anio": distribucion_anio,
            "alertas": alertas,
            "quality": {
                "report_id": str(db_quality_report.id),
                "semaforo": quality_report.get("semaforo"),
                "score": quality_report.get("score_overall"),
                "issues_count": len(quality_report.get("issues", [])),
            },
        }

        # En el caso RNMC, propagar detalles extras del ingestor (detect sheet, etc.)
        if source_id == "INSPECCION_MEDIDAS_RNMC":
            base_response.update({
                "inserted": int(res.get("inserted", 0)),
                "updated": int(res.get("updated", 0)),
                "total": int(res.get("total", 0)),
                "detected_sheet": res.get("detected_sheet"),
                "header_row": res.get("header_row"),
                "columns_detected": res.get("columns_detected"),
                "ingestor_error": res.get("error"),
                "df_shape": res.get("df_shape"),
                "municipio_uniques": res.get("municipio_uniques"),
                "detail": res.get("detail"),
            })

        return base_response

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        log_entry.estado = "FAILED"
        log_entry.errores = str(e)
        log_entry.fecha_fin = datetime.utcnow()
        db.commit()
        logger.error(f"Fallo crítico en ingestión: {e}")
        raise HTTPException(status_code=500, detail="No se pudo procesar el archivo de inteligencia.")

@router.post("/reference-upload")
async def upload_reference_file(
    request: Request,
    file: UploadFile = File(...),
    source_cutoff: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(get_optional_user),
):
    """Load aggregated nationwide reference data without altering local reports."""
    authorization_mode = await authorize_source_monitor(request, current_user, "MINDEFENSA")
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos Excel (.xlsx, .xls).")

    parsed_cutoff = None
    if source_cutoff:
        try:
            parsed_cutoff = date.fromisoformat(source_cutoff)
        except ValueError as error:
            raise HTTPException(status_code=422, detail="source_cutoff debe usar formato YYYY-MM-DD.") from error

    log_entry = IngestionLog(
        estado="IN_PROGRESS",
        registros_insertados=0,
        detalles={"filename": file.filename, "scope": "MINDEFENSA_REFERENCE"},
    )
    db.add(log_entry)
    db.commit()
    db.refresh(log_entry)

    try:
        import uuid
        from sqlalchemy.dialects.postgresql import insert

        contents = await file.read()
        file_hash = hashlib.sha256(contents).hexdigest()
        source_id = "MINDEFENSA_REFERENCE"
        existing_file = db.query(IngestionFile).filter(
            IngestionFile.file_hash == file_hash,
            IngestionFile.source_type == source_id,
        ).first()
        if existing_file:
            return {
                "ingestion_id": str(existing_file.ingestion_id),
                "status": "UNCHANGED",
                "message": "Archivo de referencia ya procesado.",
                "records": existing_file.records_count,
            }

        processor = NationalStatsProcessor()
        records = list(processor.process_reference_excel(contents, file.filename, parsed_cutoff))
        if not records:
            raise HTTPException(
                status_code=422,
                detail="El archivo no contiene municipios con codigo DANE para referencia territorial.",
            )

        period_dates = [record["fecha_hecho"] for record in records]
        years = sorted({record["anio"] for record in records})
        resolved_cutoff = records[0].get("fecha_corte_mindefensa")
        ingestion_id = uuid.uuid4()

        with db.begin_nested():
            # A countrywide monthly reference contains tens of thousands of
            # aggregates per conduct. Batched upserts avoid one database round
            # trip for every municipality-month.
            for offset in range(0, len(records), 500):
                batch = records[offset:offset + 500]
                statement = insert(NationalCrimeStats).values(batch)
                statement = statement.on_conflict_do_update(
                    index_elements=["source_id", "event_fingerprint"],
                    set_={
                        "cantidad": statement.excluded.cantidad,
                        "fecha_corte_mindefensa": statement.excluded.fecha_corte_mindefensa,
                        "fuente_archivo": statement.excluded.fuente_archivo,
                        "fecha_ingesta": statement.excluded.fecha_ingesta,
                    },
                )
                db.execute(statement)

            db.add(IngestionFile(
                ingestion_id=ingestion_id,
                filename=file.filename,
                source_type=source_id,
                file_hash=file_hash,
                inserted_count=len(records),
                updated_count=0,
                skipped_count=0,
                records_count=len(records),
                periodo_detectado=f"{min(period_dates)} a {max(period_dates)}",
                periodo_detectado_min=min(period_dates),
                periodo_detectado_max=max(period_dates),
                anios_incluidos=years,
                semanas_incluidas=[],
                status="COMPLETED",
            ))
            log_entry.estado = "COMPLETED"
            log_entry.registros_insertados = len(records)
            log_entry.fecha_fin = datetime.utcnow()
            log_entry.detalles = {
                "ingestion_id": str(ingestion_id),
                "filename": file.filename,
                "scope": source_id,
                "municipalities": len({record["codigo_dane"] for record in records}),
                "years": years,
                "source_cutoff": resolved_cutoff.isoformat() if resolved_cutoff else None,
            }

        db.commit()
        await log_audit(
            db,
            "REFERENCE_DATA_INGESTED",
            actor_id=str(current_user.id) if current_user else None,
            module="INTELLIGENCE",
            target={
                "filename": file.filename,
                "records": len(records),
                "source_cutoff": str(resolved_cutoff),
                "authorization_mode": authorization_mode,
            },
            level=2,
            request=request,
        )
        return {
            "ingestion_id": str(ingestion_id),
            "status": "COMPLETED",
            "source_id": source_id,
            "records": len(records),
            "municipalities": len({record["codigo_dane"] for record in records}),
            "years": years,
            "source_cutoff": resolved_cutoff.isoformat() if resolved_cutoff else None,
        }
    except HTTPException:
        db.rollback()
        log_entry.estado = "FAILED"
        log_entry.fecha_fin = datetime.utcnow()
        db.commit()
        raise
    except Exception as error:
        db.rollback()
        log_entry.estado = "FAILED"
        log_entry.errores = str(error)
        log_entry.fecha_fin = datetime.utcnow()
        db.commit()
        logger.exception("Reference data ingestion failed")
        raise HTTPException(status_code=500, detail="No se pudo procesar el archivo de referencia territorial.") from error


@router.post("/reference-aggregate-upload")
async def upload_reference_aggregates(
    request: Request,
    payload: ReferenceAggregateUpload,
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(get_optional_user),
):
    """Persist compact national and regional aggregates produced by the monitor.

    Raw MinDefensa workbooks are large enough to exceed a public HTTP gateway.
    The trusted monitor reads them, keeps the national total plus Valle/Cauca
    municipal aggregates, and sends the verified municipal coverage separately.
    """
    authorization_mode = await authorize_source_monitor(request, current_user, "MINDEFENSA")
    if not payload.records:
        raise HTTPException(status_code=422, detail="La referencia agregada no contiene registros.")
    if len(payload.records) > 30_000:
        raise HTTPException(status_code=413, detail="La referencia agregada excede el limite operativo.")

    log_entry = IngestionLog(
        estado="IN_PROGRESS",
        registros_insertados=0,
        detalles={"filename": payload.filename, "scope": COMPACT_REFERENCE_SOURCE},
    )
    db.add(log_entry)
    db.commit()
    db.refresh(log_entry)

    try:
        from sqlalchemy.dialects.postgresql import insert

        processor = NationalStatsProcessor()
        records = []
        for item in payload.records:
            code = "NACIONAL" if item.codigo_dane.upper() == "NACIONAL" else normalize_municipality_code(item.codigo_dane)
            if not code or not 1 <= item.mes <= 12 or item.anio < 2000 or item.cantidad < 0:
                continue
            fingerprint_source = f"{COMPACT_REFERENCE_SOURCE}|{payload.tipo_delito}|{code}|{item.anio}|{item.mes}"
            fingerprint = hashlib.sha256(fingerprint_source.encode()).hexdigest()
            records.append({
                "source_id": COMPACT_REFERENCE_SOURCE,
                "departamento": item.departamento.strip() or "NO INFORMADO",
                "municipio": item.municipio.strip() or code,
                "municipio_normalizado": processor.normalize_text(item.municipio or code),
                "codigo_dane": code,
                "fecha_hecho": date(item.anio, item.mes, 1),
                "fecha_corte_mindefensa": payload.source_cutoff,
                "anio": item.anio,
                "mes": item.mes,
                "tipo_delito": payload.tipo_delito,
                "cantidad": int(item.cantidad),
                "fuente_archivo": payload.filename,
                "event_fingerprint": fingerprint,
                "hash_registro": fingerprint,
                "fecha_ingesta": datetime.utcnow(),
            })
        if not records:
            raise HTTPException(status_code=422, detail="No hay registros agregados validos.")

        coverage = []
        for period in payload.coverage:
            codes = sorted({code for raw_code in period.municipality_codes if (code := normalize_municipality_code(raw_code))})
            if codes:
                coverage.append({
                    "source_id": COMPACT_REFERENCE_SOURCE,
                    "tipo_delito": payload.tipo_delito,
                    "anio": period.anio,
                    "municipality_codes": codes,
                    "fecha_corte_mindefensa": payload.source_cutoff,
                    "fuente_archivo": payload.filename,
                    "fecha_ingesta": datetime.utcnow(),
                })

        with db.begin_nested():
            for offset in range(0, len(records), 500):
                statement = insert(NationalCrimeStats).values(records[offset:offset + 500])
                statement = statement.on_conflict_do_update(
                    index_elements=["source_id", "event_fingerprint"],
                    set_={
                        "cantidad": statement.excluded.cantidad,
                        "fecha_corte_mindefensa": statement.excluded.fecha_corte_mindefensa,
                        "fuente_archivo": statement.excluded.fuente_archivo,
                        "fecha_ingesta": statement.excluded.fecha_ingesta,
                    },
                )
                db.execute(statement)
            for item in coverage:
                statement = insert(NationalReferenceCoverage).values(item)
                statement = statement.on_conflict_do_update(
                    index_elements=["source_id", "tipo_delito", "anio"],
                    set_={
                        "municipality_codes": statement.excluded.municipality_codes,
                        "fecha_corte_mindefensa": statement.excluded.fecha_corte_mindefensa,
                        "fuente_archivo": statement.excluded.fuente_archivo,
                        "fecha_ingesta": statement.excluded.fecha_ingesta,
                    },
                )
                db.execute(statement)
            log_entry.estado = "COMPLETED"
            log_entry.registros_insertados = len(records)
            log_entry.fecha_fin = datetime.utcnow()
            log_entry.detalles = {
                "filename": payload.filename,
                "scope": COMPACT_REFERENCE_SOURCE,
                "records": len(records),
                "coverage_years": sorted({item["anio"] for item in coverage}),
                "authorization_mode": authorization_mode,
            }
        db.commit()
        await log_audit(
            db,
            "REFERENCE_DATA_INGESTED",
            actor_id=str(current_user.id) if current_user else None,
            module="INTELLIGENCE",
            target={"filename": payload.filename, "records": len(records), "mode": "COMPACT"},
            level=2,
            request=request,
        )
        return {
            "status": "COMPLETED",
            "source_id": COMPACT_REFERENCE_SOURCE,
            "records": len(records),
            "coverage_years": sorted({item["anio"] for item in coverage}),
        }
    except HTTPException:
        db.rollback()
        log_entry.estado = "FAILED"
        log_entry.fecha_fin = datetime.utcnow()
        db.commit()
        raise
    except Exception as error:
        db.rollback()
        log_entry.estado = "FAILED"
        log_entry.errores = str(error)
        log_entry.fecha_fin = datetime.utcnow()
        db.commit()
        logger.exception("Compact reference ingestion failed")
        raise HTTPException(status_code=500, detail="No se pudo guardar la referencia territorial agregada.") from error


@router.post("/ingest")
async def trigger_ingestion(
    request: Request,
    background_tasks: BackgroundTasks, 
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["TI_ADMIN"]))
):
    """
    Inicia el proceso de descarga e ingesta de datos nacionales en segundo plano.
    """
    # Crear log de inicio
    log_entry = IngestionLog(estado="IN_PROGRESS", detalles={"trigger": "manual"})
    db.add(log_entry)
    db.commit()
    # Importante: refrescar para asegurarnos que ID existe y la DB hizo flush/commit completo
    db.refresh(log_entry)
    log_id = log_entry.id
    
    background_tasks.add_task(run_ingestion_process, log_id)
    
    return {"status": "started", "log_id": log_id, "message": "Ingesta iniciada en segundo plano"}

@router.get("/stats/compare")
async def get_crime_comparison(
    source_id: str = "SEM_POLICIA",
    type: str = "weekly",
    anio: int = None,
    semana: int = None,
    mes: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    """
    Retorna comparativo automático (WoW/MoM y YoY) para una fuente.
    """
    value = None
    if anio:
        value = {"anio": anio, "semana": semana, "mes": mes}
        
    res = IntelligenceService.get_comparison(db, source_id, type=type, value=value)
    if not res:
        return {"status": "error", "message": "No se encontraron datos para la fuente especificada"}
        
    report = IntelligenceService.format_comparison_report(res)
    return report

@router.get("/stats/ytd")
async def get_crime_ytd(
    source_id: str = "SEM_POLICIA",
    anio: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    """
    Retorna comparativo YTD (Year-To-Date) para una fuente.
    """
    res = IntelligenceService.get_ytd_comparison(db, source_id, anio=anio)
    return res

@router.get("/stats/accumulated")
async def get_crime_accumulated(
    source_id: str = "SEM_POLICIA",
    start_mm_dd: str = "01-01",
    end_mm_dd: str = "12-31",
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    """
    Acumulado de un periodo específico para todos los años disponibles.
    """
    res = IntelligenceService.get_multi_year_accumulated(db, source_id, start_mm_dd, end_mm_dd)
    return res

@router.get("/public/rnmc-summary")
async def get_public_rnmc_summary(db: Session = Depends(get_db)):
    """Public RNMC statistics. This endpoint never returns people, case files, or individual records."""
    minimum_group_size = 10
    base_filters = [
        RNMCMeasure.source_id == "INSPECCION_MEDIDAS_RNMC",
        RNMCMeasure.municipio.ilike("%JAMUNDI%"),
    ]
    latest_date = db.query(func.max(RNMCMeasure.fecha_actuacion)).filter(*base_filters).scalar()
    if not latest_date:
        return {"metadata": {"available": False, "minimum_group_size": minimum_group_size}, "kpis": {}, "monthly": [], "states": [], "measures": [], "zones": []}

    year_start = datetime(latest_date.year, 1, 1)
    year_end = datetime(latest_date.year + 1, 1, 1)
    period_filters = [*base_filters, RNMCMeasure.fecha_actuacion >= year_start, RNMCMeasure.fecha_actuacion < year_end]
    total_measures = db.query(func.count(RNMCMeasure.id)).filter(*period_filters).scalar() or 0
    total_paid = db.query(func.coalesce(func.sum(RNMCMeasure.valor_pagado), 0)).filter(*period_filters).scalar() or 0
    total_net = db.query(func.coalesce(func.sum(RNMCMeasure.valor_neto), 0)).filter(*period_filters).scalar() or 0

    def grouped(column, limit=10):
        rows = db.query(column.label("name"), func.count(RNMCMeasure.id).label("value")).filter(
            *period_filters, column.isnot(None), column != ""
        ).group_by(column).having(func.count(RNMCMeasure.id) >= minimum_group_size).order_by(func.count(RNMCMeasure.id).desc()).limit(limit).all()
        return [{"name": row.name, "value": int(row.value)} for row in rows]

    monthly_rows = db.query(
        func.extract("month", RNMCMeasure.fecha_actuacion).label("month"),
        func.count(RNMCMeasure.id).label("value"),
    ).filter(*period_filters).group_by(func.extract("month", RNMCMeasure.fecha_actuacion)).order_by(func.extract("month", RNMCMeasure.fecha_actuacion)).all()
    months = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

    return {
        "metadata": {
            "available": True,
            "year": latest_date.year,
            "cutoff": latest_date.date().isoformat(),
            "source": "Inspecciones de Policia / RNMC",
            "minimum_group_size": minimum_group_size,
            "privacy": "Datos agregados. No se publican personas, expedientes, comparendos, direcciones ni relatos.",
        },
        "kpis": {"measures": int(total_measures), "paid_value": float(total_paid), "net_value": float(total_net)},
        "monthly": [{"name": months[int(row.month) - 1], "value": int(row.value)} for row in monthly_rows],
        "states": grouped(RNMCMeasure.estado),
        "measures": grouped(RNMCMeasure.medida),
        "zones": grouped(RNMCMeasure.localidad),
    }
@router.get("/stats/rnmc")
async def get_rnmc_stats(
    request: Request,
    type: str = "monthly",
    anio: int = None,
    valor: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access)
):
    """
    Retorna estadísticas estables de RNMC para el Tab Estratégico.
    Mapea a la estructura requerida por la UI Premium.
    """
    await log_audit(
        db,
        "RNMC_STATS_VIEW",
        actor_id=str(current_user.id),
        module="RNMC",
        target={"type": type, "anio": anio},
        level=2,
        request=request
    )
    raw = RNMCService.get_rnmc_comparison(db, mode=type, anio=anio, valor=valor)
    if not raw:
        # Estructura vacía estable
        return {
            "range": {"from": None, "to": None},
            "group": type,
            "kpis": {"total": 0, "pagadas": 0, "en_proceso": 0, "ratificadas": 0, "recaudo": 0, "efectividad_pct": 0},
            "series": [],
            "by_estado": [],
            "top_medidas": [],
            "top_localidades": [],
            "comparisons": {}
        }

    actual = raw["actual"]
    series = RNMCService.get_series(db, mode="month" if type == "monthly" else "week")

    return {
        "range": actual["periodo"],
        "group": type,
        "kpis": {
            "total": actual["total_registros"],
            "pagadas": actual["pagos_conteo"],
            "en_proceso": actual.get("especificos", {}).get("en_proceso", 0),
            "ratificadas": actual.get("especificos", {}).get("ratificada", 0),
            "recaudo": actual["recaudo_total"],
            "efectividad_pct": actual.get("porcentaje_pagado", 0)
        },
        "series": series,
        "by_estado": [{"estado": k, "total": v} for k, v in actual["top_estados"].items()],
        "top_medidas": [{"medida": k, "total": v} for k, v in actual["top_medidas"].items()],
        "top_localidades": [{"localidad": k, "total": v} for k, v in actual["top_localidades"].items()],
        "comparisons": {
            "wow": raw.get("prev") if type == "weekly" else None,
            "mom": raw.get("prev") if type == "monthly" else None,
            "yoy": raw.get("yoy")
        }
    }

@router.get("/rnmc/medidas/backlog")
async def get_rnmc_backlog(
    from_date: str = None,
    to_date: str = None,
    min_dias: int = None,
    estado: str = None,
    medida: str = None,
    localidad: str = None,
    page: int = 1,
    page_size: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    """
    Lista medidas filtrables (backlog) para la UI operativa.
    """
    return RNMCService.get_backlog(
        db, from_date, to_date, min_dias, estado, medida, localidad, page, page_size
    )

@router.get("/rnmc/medidas/history")
async def get_rnmc_history(
    source_id: str = "INSPECCION_MEDIDAS_RNMC",
    event_fingerprint: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    """
    Retorna el historial de cambios de estado para una medida específica.
    """
    if not event_fingerprint:
        raise HTTPException(status_code=400, detail="event_fingerprint es requerido")
        
    res = RNMCService.get_measure_history(db, source_id, event_fingerprint)
    if not res:
        raise HTTPException(status_code=404, detail="Medida no encontrada")
    return res

class TriggerReportRequest(BaseModel):
    type: str = "all"
    source_id: str = "SEM_POLICIA"
    period: Optional[str] = None
    force: bool = False
    forced_by: Optional[str] = None
    forced_reason: Optional[str] = None

@router.post("/reports/trigger")
async def trigger_auto_reports(
    req: TriggerReportRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: Optional[User] = Depends(get_optional_user),
):
    """
    Dispara la generación de reportes automáticos (Protegido por API Key).
    Incluye soporte para periodos específicos (ej. 2026-W08).
    """
    authorization_mode = _authorize_report_trigger(request, current_user)
    trusted_actor = current_user.username if current_user else "SYSTEM_CRON"
    await log_audit(
        db,
        "AUTOMATED_REPORT_TRIGGERED",
        actor_id=str(current_user.id) if current_user else None,
        module="REPORTS",
        target={
            "type": req.type,
            "source_id": req.source_id,
            "period": req.period,
            "mode": authorization_mode,
        },
        level=2,
        request=request,
    )

    # Parsing de periodo (opcional)
    anio, valor = None, None
    if req.period:
        try:
            if "-W" in req.period:
                anio_str, valor_str = req.period.split("-W")
                anio, valor = int(anio_str), int(valor_str)
            elif "-M" in req.period:
                anio_str, valor_str = req.period.split("-M")
                anio, valor = int(anio_str), int(valor_str)
        except Exception:
            raise HTTPException(status_code=400, detail="Formato de periodo inválido. Use YYYY-WXX o YYYY-MXX")

    results = {}
    if req.source_id == "INSPECCION_MEDIDAS_RNMC":
        # Mapear types a los internos de RNMC
        if req.type in ["all", "weekly", "RNMC_WEEKLY"]:
            results["weekly"] = ReportAutomationService.run_rnmc_report(
                db, "SEMANAL", anio=anio, valor=valor, forces=req.force
            )
        if req.type in ["all", "monthly", "RNMC_MONTHLY"]:
            results["monthly"] = ReportAutomationService.run_rnmc_report(
                db, "MENSUAL", anio=anio, valor=valor, forces=req.force
            )
        if req.type in ["all", "ytd"]:
            results["ytd"] = ReportAutomationService.run_rnmc_report(
                db, "YTD", anio=anio, forces=req.force
            )
    else:
        if req.type in ["all", "weekly"]:
            results["weekly"] = ReportAutomationService.run_weekly_report(
                db, req.source_id, forces=req.force, forced_by=trusted_actor, forced_reason=req.forced_reason
            )
        if req.type in ["all", "monthly"]:
            results["monthly"] = ReportAutomationService.run_monthly_report(
                db, req.source_id, forces=req.force, forced_by=trusted_actor, forced_reason=req.forced_reason
            )
        
    return {"status": "success", "executed": {k: v.id if v else None for k, v in results.items()}, "source": req.source_id}

@router.get("/reports/history")
async def get_report_history(
    type: str = None,
    source_id: str = "SEM_POLICIA",
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    from db.models_intelligence import ReportRun
    query = db.query(ReportRun).filter(ReportRun.source_id == source_id)
    if type:
        query = query.filter(ReportRun.report_type == type.upper())
    
    reports = query.order_by(ReportRun.generated_at.desc()).limit(20).all()
    return reports

# --- ALERT FEED ENDPOINTS ---

@router.get("/alerts")
async def list_alerts(
    source: str = None,
    status: str = "OPEN",
    severity: str = None,
    tier: str = None,
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access)
):
    """
    Lista las alertas del muro de inteligencia, ordenadas por score (Fase 3).
    """
    query = db.query(IntelligenceAlert)
    
    if source:
        query = query.filter(IntelligenceAlert.source == source)
    if status:
        query = query.filter(IntelligenceAlert.status == status)
    if severity:
        query = query.filter(IntelligenceAlert.severity == severity)
    if tier:
        query = query.filter(IntelligenceAlert.priority_tier == tier)
        
    total = query.count()
    # Ordenar por action_score desc, luego por updated_at
    alerts = query.order_by(desc(IntelligenceAlert.action_score), desc(IntelligenceAlert.updated_at)).offset(offset).limit(limit).all()
    
    return {
        "total": total,
        "items": alerts,
        "limit": limit,
        "offset": offset
    }


def _build_alerts_query(
    db: Session,
    source: str,
    status: str,
    severity: Optional[str],
    tiers: Optional[list],
    from_date: Optional[str],
    to_date: Optional[str],
    limit: int,
):
    """
    Helper compartido para exportaciones y snapshots.
    Aplica filtros básicos y ordena por score descendente.
    """
    query = db.query(IntelligenceAlert).filter(IntelligenceAlert.source == source)

    if status:
        query = query.filter(IntelligenceAlert.status == status)
    if severity:
        query = query.filter(IntelligenceAlert.severity == severity)
    if tiers:
        query = query.filter(IntelligenceAlert.priority_tier.in_(tiers))

    # Filtro por rango de fechas usando created_at para trazabilidad estable
    if from_date:
        try:
            from_dt = datetime.fromisoformat(from_date)
            query = query.filter(IntelligenceAlert.created_at >= from_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Parámetro 'from' inválido. Use formato YYYY-MM-DD.")
    if to_date:
        try:
            to_dt = datetime.fromisoformat(to_date)
            query = query.filter(IntelligenceAlert.created_at <= to_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Parámetro 'to' inválido. Use formato YYYY-MM-DD.")

    capped_limit = min(max(limit, 1), 1000)

    return query.order_by(
        desc(IntelligenceAlert.action_score),
        desc(IntelligenceAlert.updated_at),
    ).limit(capped_limit)


def _sanitize_metrics(metrics: dict) -> dict:
    """
    Elimina posibles campos sensibles de métricas (PII o expediente completo).
    """
    if not metrics:
        return {}
    clean = dict(metrics)
    # Campos que nunca deben exportarse en claro
    for key in ["expediente", "documento", "identificacion", "nombre_completo"]:
        clean.pop(key, None)
    return clean


def _serialize_alert_for_export(alert: IntelligenceAlert) -> dict:
    metrics = _sanitize_metrics(alert.metrics or {})
    entity_ref = alert.entity_ref or {}

    return {
        "created_at": alert.created_at.isoformat() if alert.created_at else None,
        "updated_at": alert.updated_at.isoformat() if alert.updated_at else None,
        "source": alert.source,
        "alert_type": alert.alert_type,
        "severity": alert.severity,
        "priority_tier": alert.priority_tier,
        "action_score": float(alert.action_score or 0),
        "dias": metrics.get("dias"),
        "valor_neto": metrics.get("valor_neto"),
        "valor_pagado": metrics.get("valor_pagado"),
        "estado": metrics.get("estado"),
        "localidad": metrics.get("localidad"),
        "medida": metrics.get("medida"),
        "source_id": entity_ref.get("source_id"),
        "event_fingerprint": entity_ref.get("event_fingerprint"),
        "recommended_action": alert.recommended_action,
        "rationale_md": alert.rationale_md,
    }


@router.get("/alerts/export/excel")
async def export_alerts_excel(
    request: Request,
    source: str = "RNMC",
    status: str = "OPEN",
    severity: Optional[str] = None,
    tier: Optional[str] = Query(None, description="Tier único o lista separada por comas, ej: P1,P2"),
    from_date: Optional[str] = Query(None, alias="from"),
    to_date: Optional[str] = Query(None, alias="to"),
    limit: int = 500,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    """
    Genera y retorna un XLSX en streaming con el ranking de alertas.
    No persiste archivos en disco.
    """
    # Auditoría
    await log_audit(
        db,
        "ALERTS_EXPORT_EXCEL",
        actor_id=str(current_user.id),
        module="RNMC",
        target={"source": source, "status": status},
        level=2,
        request=request
    )
    tiers = None
    if tier:
        tiers = [t.strip() for t in tier.split(",") if t.strip()]

    query = _build_alerts_query(
        db=db,
        source=source,
        status=status,
        severity=severity,
        tiers=tiers,
        from_date=from_date,
        to_date=to_date,
        limit=limit,
    )

    alerts = query.all()
    rows = [_serialize_alert_for_export(a) for a in alerts]

    # Construir resumen
    from collections import Counter

    tier_counts = Counter(r.get("priority_tier") or "SIN_TIER" for r in rows)
    severity_counts = Counter(r.get("severity") or "SIN_SEVERIDAD" for r in rows)

    p1_unpaid_value = sum(
        float(r.get("valor_neto") or 0)
        for r in rows
        if r.get("priority_tier") == "P1" and not (r.get("valor_pagado") or 0)
    )
    total_recaudo = sum(float(r.get("valor_pagado") or 0) for r in rows)

    scoring_config = get_scoring_config()
    scoring_config_hash = hashlib.sha256(
        json.dumps(scoring_config, sort_keys=True).encode("utf-8")
    ).hexdigest()

    # Construir Excel en memoria
    from openpyxl import Workbook

    wb = Workbook()

    # Sheet 1: Ranking
    ws_rank = wb.active
    ws_rank.title = "Ranking"

    headers = [
        "created_at",
        "updated_at",
        "source",
        "alert_type",
        "severity",
        "priority_tier",
        "action_score",
        "dias",
        "valor_neto",
        "valor_pagado",
        "estado",
        "localidad",
        "medida",
        "source_id",
        "event_fingerprint",
        "recommended_action",
        "rationale_md",
    ]
    ws_rank.append(headers)
    for r in rows:
        ws_rank.append([r.get(h) for h in headers])

    # Sheet 2: Summary
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.append(["Métrica", "Valor"])
    ws_summary.append(["Total alertas", len(rows)])
    for tier_key, count in tier_counts.items():
        ws_summary.append([f"Alertas {tier_key}", count])
    for sev_key, count in severity_counts.items():
        ws_summary.append([f"Severidad {sev_key}", count])
    ws_summary.append(["Valor neto P1 sin pago", p1_unpaid_value])
    ws_summary.append(["Recaudo total (valor_pagado)", total_recaudo])

    # Sheet 3: Config
    ws_cfg = wb.create_sheet(title="Config")
    ws_cfg.append(["Clave", "Valor"])
    for key, value in scoring_config.items():
        ws_cfg.append([key, value])
    ws_cfg.append([])
    ws_cfg.append(["scoring_config_sha256", scoring_config_hash])
    ws_cfg.append(["generated_at_utc", datetime.utcnow().isoformat()])
    ws_cfg.append(["source", source])
    ws_cfg.append(["status", status])
    ws_cfg.append(["tiers", ",".join(tiers) if tiers else ""])
    ws_cfg.append(["from", from_date or ""])
    ws_cfg.append(["to", to_date or ""])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"alerts_ranking_{source}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/alerts/export/csv")
async def export_alerts_csv(
    request: Request,
    source: str = "RNMC",
    status: str = "OPEN",
    severity: Optional[str] = None,
    tier: Optional[str] = Query(None, description="Tier único o lista separada por comas, ej: P1,P2"),
    from_date: Optional[str] = Query(None, alias="from"),
    to_date: Optional[str] = Query(None, alias="to"),
    limit: int = 500,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    """
    Exportación rápida en CSV del ranking de alertas.
    """
    # Auditoría
    await log_audit(
        db,
        "ALERTS_EXPORT_CSV",
        actor_id=str(current_user.id),
        module="RNMC",
        target={"source": source, "status": status},
        level=2,
        request=request
    )
    tiers = None
    if tier:
        tiers = [t.strip() for t in tier.split(",") if t.strip()]

    query = _build_alerts_query(
        db=db,
        source=source,
        status=status,
        severity=severity,
        tiers=tiers,
        from_date=from_date,
        to_date=to_date,
        limit=limit,
    )
    alerts = query.all()
    rows = [_serialize_alert_for_export(a) for a in alerts]

    headers = [
        "created_at",
        "updated_at",
        "source",
        "alert_type",
        "severity",
        "priority_tier",
        "action_score",
        "dias",
        "valor_neto",
        "valor_pagado",
        "estado",
        "localidad",
        "medida",
        "source_id",
        "event_fingerprint",
        "recommended_action",
        "rationale_md",
    ]

    import csv

    stream = BytesIO()
    writer = csv.writer(stream)
    writer.writerow(headers)
    for r in rows:
        writer.writerow([r.get(h, "") for h in headers])
    stream.seek(0)

    filename = f"alerts_ranking_{source}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        stream,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

async def process_alert_ai(alert_id: UUID, db: Session):
    """Tarea en background para generar el rationale de la IA."""
    alert = db.query(IntelligenceAlert).filter(IntelligenceAlert.id == alert_id).first()
    if not alert: return
    
    # Preparamos el input del score (ya calculado)
    scoring_data = {
        "action_score": float(alert.action_score or 0),
        "priority_tier": alert.priority_tier
    }
    
    ai_res = await build_ai_rationale(alert, scoring_data)
    if ai_res.get("ai_rationale_md"):
        alert.ai_rationale_md = ai_res["ai_rationale_md"]
        alert.ai_provider = ai_res["ai_provider"]
        alert.ai_request_id = ai_res["ai_request_id"]
        db.commit()


class AlertsSnapshotRequest(BaseModel):
    source: str = "RNMC"
    status: str = "OPEN"
    tiers: Optional[List[str]] = None
    severity: Optional[str] = None
    from_date: Optional[str] = None
    to_date: Optional[str] = None
    limit: int = 500


def _create_snapshot_from_rows(
    db: Session,
    source: str,
    filters: dict,
    rows: List[dict],
    scoring_config: dict,
):
    """
    Crea (o reutiliza) un snapshot inmutable a partir de filas ya serializadas.
    """
    payload = {"alerts": rows}
    payload_str = json.dumps(payload, sort_keys=True, default=str)
    payload_sha256 = hashlib.sha256(payload_str.encode("utf-8")).hexdigest()

    # Intentar reutilizar snapshot por hash (idempotencia)
    existing = db.query(IntelligenceAlertSnapshot).filter(
        IntelligenceAlertSnapshot.payload_sha256 == payload_sha256
    ).first()
    if existing:
        return existing

    snapshot = IntelligenceAlertSnapshot(
        source=source,
        filters=filters,
        scoring_config=scoring_config,
        payload_json=payload,
        payload_sha256=payload_sha256,
    )
    db.add(snapshot)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        # Carrera poco probable: recuperar el existente por hash
        snapshot = db.query(IntelligenceAlertSnapshot).filter(
            IntelligenceAlertSnapshot.payload_sha256 == payload_sha256
        ).first()
    db.refresh(snapshot)
    return snapshot


def _build_alerts_pdf_html(
    rows: List[dict],
    scoring_config: dict,
    snapshot_id: str,
    payload_sha256: str,
    filters: dict,
):
    """
    Construye HTML para el PDF ejecutivo del ranking de alertas RNMC.
    """
    total = len(rows)
    p1 = [r for r in rows if r.get("priority_tier") == "P1"]
    p2 = [r for r in rows if r.get("priority_tier") == "P2"]
    p3 = [r for r in rows if r.get("priority_tier") == "P3"]

    top_p1 = sorted(p1, key=lambda r: float(r.get("action_score") or 0), reverse=True)[:10]

    valor_neto_p1_sin_pago = sum(
        float(r.get("valor_neto") or 0)
        for r in p1
        if not (r.get("valor_pagado") or 0)
    )

    from collections import Counter

    localidades_p1 = Counter(r.get("localidad") or "SIN_LOCALIDAD" for r in p1)
    top_localidades = localidades_p1.most_common(5)

    period_label = "Sin rango definido"
    if filters.get("from") or filters.get("to"):
        period_label = f"{filters.get('from') or '...'} a {filters.get('to') or '...'}"

    scoring_config_hash = hashlib.sha256(
        json.dumps(scoring_config, sort_keys=True).encode("utf-8")
    ).hexdigest()

    # HTML simple de 1–2 páginas
    rows_html = ""
    for r in top_p1:
        rows_html += f"""
        <tr>
            <td>{r.get('localidad') or '-'}</td>
            <td>{r.get('medida') or '-'}</td>
            <td>{r.get('dias') or '-'}</td>
            <td>{r.get('valor_neto') or 0:,.0f}</td>
            <td>{r.get('action_score') or 0:.2f}</td>
            <td>{(r.get('recommended_action') or '').split('.')[0]}</td>
        </tr>
        """

    loc_rows = ""
    for loc, cnt in top_localidades:
        loc_rows += f"""
        <tr>
            <td>{loc}</td>
            <td>{cnt}</td>
        </tr>
        """

    html = f"""
    <html>
    <head>
        <meta charset="utf-8" />
        <title>Ranking RNMC - SISC Jamundí</title>
        <style>
            @page {{
                size: A4;
                margin: 2cm;
            }}
            body {{
                font-family: 'Helvetica', 'Arial', sans-serif;
                color: #0f172a;
                line-height: 1.5;
                font-size: 11pt;
            }}
            h1, h2, h3 {{
                color: #0f172a;
            }}
            .header {{
                border-bottom: 2px solid #0f172a;
                padding-bottom: 8px;
                margin-bottom: 18px;
            }}
            .kpi-grid {{
                display: flex;
                gap: 16px;
                margin-bottom: 18px;
            }}
            .card {{
                flex: 1;
                border-radius: 8px;
                border: 1px solid #e2e8f0;
                padding: 10px 12px;
                background: #f8fafc;
            }}
            .card-label {{
                font-size: 9pt;
                text-transform: uppercase;
                color: #64748b;
                font-weight: bold;
            }}
            .card-value {{
                font-size: 18pt;
                font-weight: bold;
                color: #0f172a;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 10px;
                margin-bottom: 10px;
            }}
            th, td {{
                border: 1px solid #e2e8f0;
                padding: 6px 8px;
                font-size: 9pt;
            }}
            th {{
                background-color: #f1f5f9;
                text-align: left;
            }}
            .audit-box {{
                margin-top: 18px;
                padding: 10px 12px;
                border-radius: 8px;
                border: 1px solid #e2e8f0;
                background: #f9fafb;
                font-size: 8.5pt;
            }}
            .audit-row {{
                display: flex;
                justify-content: space-between;
                margin-bottom: 4px;
            }}
            .audit-label {{
                color: #64748b;
                font-weight: bold;
            }}
            .audit-value {{
                color: #0f172a;
                font-family: monospace;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>SISC Jamundí — Ranking Ejecutivo RNMC</h1>
            <div style="font-size: 10pt; color: #64748b;">
                Fuente: RNMC · Periodo: {period_label} · Generado: {datetime.utcnow().isoformat()} (UTC)
            </div>
        </div>

        <div class="kpi-grid">
            <div class="card">
                <div class="card-label">Total alertas consideradas</div>
                <div class="card-value">{total}</div>
            </div>
            <div class="card">
                <div class="card-label">Prioridad Inmediata (P1)</div>
                <div class="card-value">{len(p1)}</div>
            </div>
            <div class="card">
                <div class="card-label">Gestión semanal (P2)</div>
                <div class="card-value">{len(p2)}</div>
            </div>
            <div class="card">
                <div class="card-label">Monitoreo (P3)</div>
                <div class="card-value">{len(p3)}</div>
            </div>
        </div>

        <h2>Top 10 alertas P1 por score</h2>
        <table>
            <thead>
                <tr>
                    <th>Localidad</th>
                    <th>Medida</th>
                    <th>Días</th>
                    <th>Valor Neto</th>
                    <th>Score</th>
                    <th>Razón Corta</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>

        <h2>Resumen de cartera prioritaria</h2>
        <p>Monto neto estimado sin pago en Tier P1: <strong>${valor_neto_p1_sin_pago:,.0f}</strong></p>

        <h3>Top localidades por número de P1</h3>
        <table>
            <thead>
                <tr>
                    <th>Localidad</th>
                    <th>Alertas P1</th>
                </tr>
            </thead>
            <tbody>
                {loc_rows}
            </tbody>
        </table>

        <div class="audit-box">
            <div style="font-weight: bold; text-transform: uppercase; margin-bottom: 6px;">Sello de evidencia digital</div>
            <div class="audit-row">
                <span class="audit-label">Snapshot ID</span>
                <span class="audit-value">{snapshot_id}</span>
            </div>
            <div class="audit-row">
                <span class="audit-label">Payload SHA256</span>
                <span class="audit-value">{payload_sha256}</span>
            </div>
            <div class="audit-row">
                <span class="audit-label">Scoring config hash</span>
                <span class="audit-value">{scoring_config_hash}</span>
            </div>
        </div>
    </body>
    </html>
    """
    return html


class AlertsPdfExportRequest(BaseModel):
    source: str = "RNMC"
    status: str = "OPEN"
    tiers: Optional[List[str]] = None
    severity: Optional[str] = None
    from_date: Optional[str] = None
    to_date: Optional[str] = None
    limit: int = 500
    snapshot_id: Optional[UUID] = None

@router.post("/alerts/prioritize")
async def prioritize_alerts(
    source: str = "RNMC",
    ai: bool = True,
    bg_tasks: BackgroundTasks = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access)
):
    """
    Fase 3: Ejecuta el scoring de prioridad para todas las alertas OPEN de una fuente.
    """
    alerts = db.query(IntelligenceAlert).filter(
        IntelligenceAlert.source == source,
        IntelligenceAlert.status == "OPEN"
    ).all()
    
    count = 0
    for alert in alerts:
        score_res = compute_action_score(alert)
        for k, v in score_res.items():
            setattr(alert, k, v)
        
        if ai and bg_tasks:
            bg_tasks.add_task(process_alert_ai, alert.id, db)
        
        count += 1
    
    db.commit()
    return {"status": "success", "prioritized_count": count, "ai_tasks_queued": ai}


@router.post("/alerts/snapshot")
async def create_alerts_snapshot(
    req: AlertsSnapshotRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    """
    Crea un snapshot inmutable del ranking de alertas (sin PII) y lo persiste en DB.
    """
    query = _build_alerts_query(
        db=db,
        source=req.source,
        status=req.status,
        severity=req.severity,
        tiers=req.tiers,
        from_date=req.from_date,
        to_date=req.to_date,
        limit=req.limit,
    )
    alerts = query.all()
    rows = [_serialize_alert_for_export(a) for a in alerts]
    scoring_config = get_scoring_config()

    filters = {
        "source": req.source,
        "status": req.status,
        "tiers": req.tiers,
        "severity": req.severity,
        "from": req.from_date,
        "to": req.to_date,
        "limit": req.limit,
    }

    snapshot = _create_snapshot_from_rows(
        db=db,
        source=req.source,
        filters=filters,
        rows=rows,
        scoring_config=scoring_config,
    )

    return {
        "snapshot_id": str(snapshot.id),
        "sha256": snapshot.payload_sha256,
        "created_at": snapshot.created_at,
        "source": snapshot.source,
    }

@router.get("/alerts/{alert_id}")
async def get_alert_detail(
    alert_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access)
):
    alert = db.query(IntelligenceAlert).filter(IntelligenceAlert.id == alert_id).first()
    if not alert:
        raise HTTPException(status_code=404, detail="Alerta no encontrada")
    return alert

@router.post("/alerts/{alert_id}/ack")
async def acknowledge_alert(
    alert_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access)
):
    alert = db.query(IntelligenceAlert).filter(IntelligenceAlert.id == alert_id).first()
    if not alert:
        raise HTTPException(status_code=404, detail="Alerta no encontrada")
    
    alert.status = "ACK"
    db.commit()
    return {"status": "success", "message": "Alerta marcada como reconocida"}

@router.post("/alerts/{alert_id}/dismiss")
async def dismiss_alert(
    alert_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access)
):
    alert = db.query(IntelligenceAlert).filter(IntelligenceAlert.id == alert_id).first()
    if not alert:
        raise HTTPException(status_code=404, detail="Alerta no encontrada")
    
    alert.status = "DISMISSED"
    db.commit()
    return {"status": "success", "message": "Alerta descartada"}

@router.post("/alerts/rnmc/generate")
async def trigger_rnmc_alerts(
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access)
):
    """
    Genera alertas manualmente para el módulo RNMC.
    """
    res = generate_rnmc_alerts(db)
    return res


@router.post("/alerts/export/pdf")
async def export_alerts_pdf(
    req: AlertsPdfExportRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    """
    Genera un PDF ejecutivo on-demand para el ranking de alertas.
    Si no se proporciona snapshot_id, crea un snapshot nuevo y lo usa como evidencia.
    """
    # Auditoría
    await log_audit(
        db,
        "ALERTS_EXPORT_PDF",
        actor_id=str(current_user.id),
        module="RNMC",
        target={"source": req.source, "snapshot_id": str(req.snapshot_id) if req.snapshot_id else "NEW"},
        level=2,
        request=request
    )
    if req.snapshot_id:
        snapshot = db.query(IntelligenceAlertSnapshot).filter(
            IntelligenceAlertSnapshot.id == req.snapshot_id
        ).first()
        if not snapshot:
            raise HTTPException(status_code=404, detail="Snapshot no encontrado")

        scoring_config = snapshot.scoring_config or get_scoring_config()
        payload = snapshot.payload_json or {}
        rows = payload.get("alerts", [])
        filters = snapshot.filters or {}
        snapshot_id_str = str(snapshot.id)
        payload_sha256 = snapshot.payload_sha256
    else:
        query = _build_alerts_query(
            db=db,
            source=req.source,
            status=req.status,
            severity=req.severity,
            tiers=req.tiers,
            from_date=req.from_date,
            to_date=req.to_date,
            limit=req.limit,
        )
        alerts = query.all()
        rows = [_serialize_alert_for_export(a) for a in alerts]
        scoring_config = get_scoring_config()
        filters = {
            "source": req.source,
            "status": req.status,
            "tiers": req.tiers,
            "severity": req.severity,
            "from": req.from_date,
            "to": req.to_date,
            "limit": req.limit,
        }
        snapshot = _create_snapshot_from_rows(
            db=db,
            source=req.source,
            filters=filters,
            rows=rows,
            scoring_config=scoring_config,
        )
        snapshot_id_str = str(snapshot.id)
        payload_sha256 = snapshot.payload_sha256

    html = _build_alerts_pdf_html(
        rows=rows,
        scoring_config=scoring_config,
        snapshot_id=snapshot_id_str,
        payload_sha256=payload_sha256,
        filters=filters,
    )

    css = CSS(
        string="""
        @page { size: A4; margin: 2cm; }
    """
    )
    pdf_io = BytesIO()
    HTML(string=html).write_pdf(pdf_io, stylesheets=[css])
    pdf_io.seek(0)

    filename = f"alerts_ranking_{filters.get('source', 'RNMC')}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        pdf_io,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/alerts/scoring-config")
async def get_alerts_scoring_config(
    current_user: User = Depends(institutional_access),
):
    """
    Devuelve la configuración actual de scoring (pesos y umbrales) para transparencia en UI.
    """
    return get_scoring_config()

@router.get("/reports/{report_run_id}")
async def get_report_detail(
    report_run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    report = db.query(ReportRun).filter(ReportRun.id == report_run_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Reporte no encontrado")
    return report

@router.post("/reports/{report_run_id}/export/pdf")
async def generate_report_pdf(
    report_run_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access)
):
    """
    Genera PDF si no existe y devuelve metadatos + enlace seguro.
    """
    report = db.query(ReportRun).filter(ReportRun.id == report_run_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Reporte no encontrado")
    
    # Generar PDF si no existe
    pdf_path = PdfReportService.generate_pdf(db, report)
    
    # Generar token de descarga de un solo uso
    token_obj = DistributionService.generate_secure_token(db, report_run_id)
    base_url = str(request.base_url).rstrip('/')
    secure_link = f"{base_url}/api/intelligence/reports/{report_run_id}/export/pdf?token={token_obj.token}"
    
    # Convertir generated_at a hora de Bogotá (UTC-5)
    from datetime import timedelta
    generated_at_bogota = (report.pdf_generated_at - timedelta(hours=5)).isoformat() if report.pdf_generated_at else None
    
    return {
        "status": "success",
        "report_run_id": report.id,
        "report_type": report.report_type,
        "period_key": report.period_key,
        "generated_at_bogota": generated_at_bogota,
        "pdf_sha256": report.pdf_sha256,
        "download_link": secure_link
    }


@router.get("/reports/{report_run_id}/export/pdf")
async def download_report_pdf(
    report_run_id: int,
    token: str = None,
    db: Session = Depends(get_db),
    request: Request = None,
    user: Optional[User] = Depends(get_optional_user) # Nueva dependencia opcional
):
    """
    Descarga segura de PDF: requiere JWT o un Token válido.
    Genera marca de agua personalizada por descarga.
    """
    report = db.query(ReportRun).filter(ReportRun.id == report_run_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Reporte no encontrado")
    
    token_db = None
    user_id = None
    user_name = "INVITADO"
    
    # 1. Validación por Token Externo
    if token:
        from db.models_intelligence import ReportDownloadToken
        token_db = db.query(ReportDownloadToken).filter(
            ReportDownloadToken.token == token,
            ReportDownloadToken.report_run_id == report_run_id,
            ReportDownloadToken.expires_at > datetime.utcnow(),
            ReportDownloadToken.is_used == False
        ).first()
        
    # 2. Validación por Sesión si no hay token
    if not token_db:
        if not user:
            raise HTTPException(status_code=401, detail="Se requiere autenticación o token válido")
        user_id = user.id
        user_name = user.full_name or user.username
    else:
        user_name = f"PORTAL_EXTERNO_{token_db.token[:6]}"

    # 3. Auditoría Institucional
    await log_audit(
        db, 
        "PDF_EXPORT", 
        actor_id=str(user_id) if user_id else None, 
        module="REPORTS",
        target={"report_id": report_run_id, "period": report.period_key},
        level=2,
        request=request
    )
    
    # 4. Generar PDF bajo demanda con Marca de Agua
    # Esto asegura que el PDF servido tenga el nombre del usuario que lo descarga
    pdf_path = PdfReportService.generate_pdf(db, report, user_name=user_name)
    
    if token_db:
        token_db.is_used = True
        db.commit()

    return FileResponse(
        pdf_path,
        media_type='application/pdf',
        filename=os.path.basename(pdf_path)
    )

@router.post("/reports/{report_run_id}/notify")
async def notify_report_results(
    report_run_id: int,
    group: str,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["TI_ADMIN", "FUNC_ADMIN", "ANALYST"]))
):
    """
    Dispara la notificación manual/automática a un grupo.
    """
    # Auditoría
    await log_audit(
        db,
        "REPORT_NOTIFY",
        actor_id=str(current_user.id),
        module="REPORTS",
        target={"report_id": report_run_id, "group": group},
        level=2,
        request=request
    )
    # 1. Asegurar que existe PDF
    report = db.query(ReportRun).filter(ReportRun.id == report_run_id).first()
    if not report or not report.pdf_path:
         raise HTTPException(status_code=400, detail="Debe generar el PDF antes de notificar.")
    
    base_url = str(request.base_url).rstrip("/")
    # Generar un token para el link de distribución
    token_obj = DistributionService.generate_secure_token(db, report_run_id)
    secure_link = f"{base_url}/api/intelligence/reports/{report_run_id}/export/pdf?token={token_obj.token}"
    
    result = DistributionService.notify_group(db, report_run_id, group, secure_link)
    result["secure_link_generated"] = secure_link
    return result

@router.get("/ingest/status/{log_id}")
async def get_ingestion_status(
    log_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    """
    Retorna el estado de una tarea de ingesta en background.
    """
    log_entry = db.query(IngestionLog).filter(IngestionLog.id == log_id).first()
    if not log_entry:
        raise HTTPException(status_code=404, detail="Log_id no encontrado")

    return {
        "id": log_entry.id,
        "estado": log_entry.estado,
        "registros_insertados": log_entry.registros_insertados,
        "archivos_procesados": log_entry.archivos_procesados,
        "errores": log_entry.errores,
        "fecha_inicio": log_entry.fecha_inicio,
        "fecha_fin": log_entry.fecha_fin,
        "detalles": log_entry.detalles
    }

def run_ingestion_process(log_id: int):
    try:
        from services.scraper_mindefensa import MinDefensaScraper
        from services.scraper_policia import PoliciaScraper
        from services.excel_processor import NationalStatsProcessor
        from db.models import SessionLocal
        
        db_bg = SessionLocal()
        
        # Wait a moment to ensure the web request transaction is fully visible
        import time
        time.sleep(1)
        
        log = db_bg.query(IngestionLog).filter(IngestionLog.id == log_id).first()
        if not log:
            logger.error(f"Error crítico: log_id {log_id} no encontrado en background task.")
            db_bg.close()
            return
            
        log.estado = 'IN_PROGRESS'
        db_bg.commit()
        
        scraper_mindefensa = MinDefensaScraper()
        scraper_policia = PoliciaScraper()
        processor = NationalStatsProcessor()
        
        # 1. Obtener archivos ya procesados exitosamente en el pasado
        processed_files_names = set()
        past_successful_logs = db_bg.query(IngestionLog).filter(
            IngestionLog.estado == 'SUCCESS',
            IngestionLog.id < log_id
        ).all()
        for p_log in past_successful_logs:
            if p_log.detalles and "processed_file_list" in p_log.detalles:
                processed_files_names.update(p_log.detalles["processed_file_list"])

        # 2. Combinar listas de archivos de ambas fuentes
        files_md = scraper_mindefensa.fetch_available_files()
        try:
            files_policia = scraper_policia.fetch_available_files()
        except Exception as e:
            logger.warning(f"No se pudieron obtener archivos de la Policía: {e}")
            files_policia = []
            
        all_remote_files = files_md + files_policia
        
        # 3. Filtrado inteligente
        current_year = datetime.now().year
        files_to_process = []
        skipped_files = []
        
        for f in all_remote_files:
            file_year = f.get('year', 2025)
            # Solo omitir si es un año pasado Y ya fue procesado con éxito
            if file_year < current_year and f['name'] in processed_files_names:
                skipped_files.append(f['name'])
            else:
                files_to_process.append(f)

        total_files = len(files_to_process)
        log.detalles = {
            "found_files": len(all_remote_files),
            "files_to_process": total_files,
            "skipped_count": len(skipped_files),
            "skipped_files": skipped_files,
            "processed_file_list": [] # Se llenará conforme se procesen
        }
        db_bg.commit()
        
        records_inserted = 0
        processed_count = 0
        total_inserted = 0
        
        processed_file_list = []
        
        for file_info in files_to_process:
            # ACTUALIZAR PROGRESO AL INICIO DE CADA ARCHIVO
            import copy
            from sqlalchemy.orm.attributes import flag_modified
            
            new_detalles = copy.deepcopy(log.detalles) if log.detalles else {}
            new_detalles["current_file"] = file_info['name']
            new_detalles["processed_files"] = processed_count 
            new_detalles["progress"] = round((processed_count / total_files) * 100) if total_files > 0 else 0
            
            log.detalles = new_detalles
            flag_modified(log, "detalles")
            db_bg.commit()
            
            logger.info(f"Iniciando procesamiento de {file_info['name']} (Progreso: {new_detalles['progress']}%)")

            # Avanzar contador para la UI
            processed_count += 1

            try:
                # Seleccionar scraper adecuado vía URL
                if 'policia.gov.co' in file_info['url']:
                    content = scraper_policia.download_file(file_info['url'])
                else:
                    content = scraper_mindefensa.download_file(file_info['url'])
                    
                if content:
                    records_generator = processor.process_excel(content, file_info['name'])
                    
                    batch = []
                    BATCH_SIZE = 500 # Un poco más conservador para evitar OOM
                    from sqlalchemy.dialects.postgresql import insert

                    for record_dict in records_generator:
                        batch.append(record_dict)
                        
                        if len(batch) >= BATCH_SIZE:
                            try:
                                stmt = insert(NationalCrimeStats).values(batch)
                                stmt = stmt.on_conflict_do_nothing(index_elements=['hash_registro'])
                                db_bg.execute(stmt)
                                db_bg.commit()
                                total_inserted += len(batch)
                            except Exception as batch_err:
                                db_bg.rollback()
                                logger.error(f"Error en bloque de {file_info['name']}: {batch_err}")
                                # Inserción individual si el bloque falla (duplicados, etc)
                                for r in batch:
                                    try:
                                        db_bg.add(NationalCrimeStats(**r))
                                        db_bg.commit()
                                        total_inserted += 1
                                    except Exception:
                                        db_bg.rollback()
                                        continue
                            batch = []
                            import gc
                            gc.collect()

                    # Guardar remanente
                    if batch:
                        try:
                            stmt = insert(NationalCrimeStats).values(batch)
                            stmt = stmt.on_conflict_do_nothing(index_elements=['hash_registro'])
                            db_bg.execute(stmt)
                            db_bg.commit()
                            total_inserted += len(batch)
                        except Exception as rem_err:
                            db_bg.rollback()
                            for r in batch:
                                try:
                                    db_bg.add(NationalCrimeStats(**r))
                                    db_bg.commit()
                                    total_inserted += 1
                                except Exception:
                                    db_bg.rollback()
                                    continue
                    
                # Liberar memoria
                if 'content' in locals(): del content
                import gc
                gc.collect()

                # Registro de éxito para este archivo
                processed_file_list.append(file_info['name'])
                log.detalles["processed_file_list"] = processed_file_list
                flag_modified(log, "detalles")
                db_bg.commit()

            except Exception as loop_err:
                logger.error(f"Error inesperado procesando archivo {file_info['name']}: {loop_err}")
            
        log.estado = "SUCCESS"
        log.archivos_procesados = processed_count
        log.registros_insertados = total_inserted
        log.fecha_fin = datetime.utcnow()
        db_bg.commit()
        
    except Exception as e:
        logger.error(f"Error crítico en ingesta background: {e}")
        if log:
            log.estado = "ERROR"
            log.errores = str(e)
            log.fecha_fin = datetime.utcnow()
            db_bg.commit()
    finally:
        db_bg.close()

@router.get("/executive-brief")
async def get_executive_brief(
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    """
    Retorna un resumen ejecutivo ágil con IA y fechas de corte para Jamundí.
    """
    from sqlalchemy import func
    from db.models_intelligence import NationalCrimeStats
    
    count_jamundi = db.query(func.count(NationalCrimeStats.id)).filter(
        NationalCrimeStats.municipio_normalizado.ilike('%JAMUNDI%')
    ).scalar()
    
    delitos_disponibles = db.query(NationalCrimeStats.tipo_delito).filter(
        NationalCrimeStats.municipio_normalizado.ilike('%JAMUNDI%')
    ).distinct().all()
    
    briefs = await IntelligenceService.get_executive_brief(db)
    
    return {
        "debug": {
            "total_jamundi": count_jamundi,
            "delitos": [d[0] for d in delitos_disponibles]
        },
        "briefs": briefs
    }

@router.get("/stats")
async def get_national_stats(
    municipio: str = "JAMUNDI",
    anio: int = 2025,
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    """
    Retorna contexto histórico local de MinDefensa con comparación anual equivalente.
    """
    from sqlalchemy import func
    
    # 1. Normalizar municipio
    processor = NationalStatsProcessor()
    target_municipio = processor.normalize_text(municipio)

    # Prefer compact aggregates from the trusted monitor. The legacy full-file
    # route remains a fallback for historical data already ingested.
    compact_source = NationalCrimeStats.source_id == COMPACT_REFERENCE_SOURCE
    has_compact_reference = db.query(NationalCrimeStats.id).filter(
        compact_source,
        NationalCrimeStats.municipio_normalizado == target_municipio,
        NationalCrimeStats.anio == anio,
    ).first() is not None
    reference_source = NationalCrimeStats.source_id == "MINDEFENSA_REFERENCE"
    has_reference_data = db.query(NationalCrimeStats.id).filter(
        reference_source,
        NationalCrimeStats.municipio_normalizado == target_municipio,
        NationalCrimeStats.anio == anio,
    ).first() is not None
    reference_source_id = (
        COMPACT_REFERENCE_SOURCE if has_compact_reference
        else "MINDEFENSA_REFERENCE" if has_reference_data
        else None
    )
    mindefensa_source = (
        NationalCrimeStats.source_id == reference_source_id
        if reference_source_id else NationalCrimeStats.source_id.ilike("%MINDEFENSA%")
    )
    is_compact_reference = reference_source_id == COMPACT_REFERENCE_SOURCE
    
    # 2. Obtener datos locales (Jamundí o el seleccionado)
    local_data = db.query(
        NationalCrimeStats.tipo_delito,
        func.sum(NationalCrimeStats.cantidad).label("total"),
        func.max(NationalCrimeStats.mes).label("period_end_month"),
    ).filter(
        mindefensa_source,
        NationalCrimeStats.municipio_normalizado == target_municipio,
        NationalCrimeStats.anio == anio
    ).group_by(NationalCrimeStats.tipo_delito).all()

    # 3. Tendencia mensual local
    trend_data = db.query(
        NationalCrimeStats.mes,
        func.sum(NationalCrimeStats.cantidad).label("total")
    ).filter(
        mindefensa_source,
        NationalCrimeStats.municipio_normalizado == target_municipio,
        NationalCrimeStats.anio == anio
    ).group_by(NationalCrimeStats.mes).order_by(NationalCrimeStats.mes).all()

    # 4. LÓGICA DE COMPARATIVA (YoY)
    # Obtener totales del año anterior (mismo municipio)
    yoy_data = db.query(
        NationalCrimeStats.tipo_delito,
        NationalCrimeStats.mes,
        func.sum(NationalCrimeStats.cantidad).label("total")
    ).filter(
        mindefensa_source,
        NationalCrimeStats.municipio_normalizado == target_municipio,
        NationalCrimeStats.anio == anio - 1
    ).group_by(NationalCrimeStats.tipo_delito, NationalCrimeStats.mes).all()
    yoy_monthly = {}
    for yoy_row in yoy_data:
        yoy_monthly.setdefault(yoy_row.tipo_delito, {})[int(yoy_row.mes)] = int(yoy_row.total)

    local_code_rows = db.query(NationalCrimeStats.codigo_dane).filter(
        mindefensa_source,
        NationalCrimeStats.municipio_normalizado == target_municipio,
        NationalCrimeStats.anio == anio,
        NationalCrimeStats.codigo_dane.isnot(None),
    ).distinct().all()
    local_code = next(
        (normalize_municipality_code(row.codigo_dane) for row in local_code_rows if normalize_municipality_code(row.codigo_dane)),
        None,
    ) or municipality_code_for_name(municipio, anio)
    territorial_peer_codes = population_peer_codes(local_code, anio)

    coverage_codes_by_type = {}
    national_totals_by_type = {}
    territorial_totals_by_type = {}
    territorial_totals_by_code_by_type = {}
    territorial_codes_by_type = {}
    cutoffs_by_type = {}
    all_source_codes = set()
    if is_compact_reference:
        national_rows = db.query(
            NationalCrimeStats.tipo_delito,
            func.sum(NationalCrimeStats.cantidad).label("total"),
        ).filter(
            mindefensa_source,
            NationalCrimeStats.anio == anio,
            NationalCrimeStats.codigo_dane == "NACIONAL",
        ).group_by(NationalCrimeStats.tipo_delito).all()
        territorial_rows = db.query(
            NationalCrimeStats.tipo_delito,
            NationalCrimeStats.codigo_dane,
            func.sum(NationalCrimeStats.cantidad).label("total"),
        ).filter(
            mindefensa_source,
            NationalCrimeStats.anio == anio,
            NationalCrimeStats.codigo_dane.in_(territorial_peer_codes),
        ).group_by(NationalCrimeStats.tipo_delito, NationalCrimeStats.codigo_dane).all()
        coverage_rows = db.query(NationalReferenceCoverage).filter(
            NationalReferenceCoverage.source_id == COMPACT_REFERENCE_SOURCE,
            NationalReferenceCoverage.anio == anio,
        ).all()
        for national_row in national_rows:
            national_totals_by_type[national_row.tipo_delito] = int(national_row.total)
        for territorial_row in territorial_rows:
            code = normalize_municipality_code(territorial_row.codigo_dane)
            if not code:
                continue
            territorial_totals_by_type[territorial_row.tipo_delito] = (
                territorial_totals_by_type.get(territorial_row.tipo_delito, 0) + int(territorial_row.total)
            )
            territorial_totals_by_code_by_type.setdefault(territorial_row.tipo_delito, {})[code] = int(territorial_row.total)
            territorial_codes_by_type.setdefault(territorial_row.tipo_delito, set()).add(code)
        for coverage_row in coverage_rows:
            codes = {
                code for raw_code in (coverage_row.municipality_codes or [])
                if (code := normalize_municipality_code(raw_code))
            }
            coverage_codes_by_type[coverage_row.tipo_delito] = codes
            all_source_codes.update(codes)
            if coverage_row.fecha_corte_mindefensa:
                cutoffs_by_type.setdefault(coverage_row.tipo_delito, set()).add(
                    coverage_row.fecha_corte_mindefensa
                )
    else:
        national_rows = db.query(
            NationalCrimeStats.tipo_delito,
            NationalCrimeStats.codigo_dane,
            func.sum(NationalCrimeStats.cantidad).label("total"),
        ).filter(
            mindefensa_source,
            NationalCrimeStats.anio == anio,
            NationalCrimeStats.codigo_dane.isnot(None),
            ~NationalCrimeStats.municipio_normalizado.like("TOTAL%"),
        ).group_by(
            NationalCrimeStats.tipo_delito,
            NationalCrimeStats.codigo_dane,
        ).all()
        cutoff_rows = db.query(
            NationalCrimeStats.tipo_delito,
            NationalCrimeStats.fecha_corte_mindefensa,
        ).filter(
            mindefensa_source,
            NationalCrimeStats.anio == anio,
            NationalCrimeStats.fecha_corte_mindefensa.isnot(None),
        ).distinct().all()
        for national_row in national_rows:
            source_code = normalize_municipality_code(national_row.codigo_dane)
            if not source_code:
                continue
            coverage_codes_by_type.setdefault(national_row.tipo_delito, set()).add(source_code)
            national_totals_by_type[national_row.tipo_delito] = (
                national_totals_by_type.get(national_row.tipo_delito, 0) + int(national_row.total)
            )
            if source_code in territorial_peer_codes:
                territorial_totals_by_type[national_row.tipo_delito] = (
                    territorial_totals_by_type.get(national_row.tipo_delito, 0) + int(national_row.total)
                )
                territorial_codes_by_type.setdefault(national_row.tipo_delito, set()).add(source_code)
                territorial_totals_by_code_by_type.setdefault(national_row.tipo_delito, {})[source_code] = int(national_row.total)
            all_source_codes.add(source_code)
        for cutoff_row in cutoff_rows:
            cutoffs_by_type.setdefault(cutoff_row.tipo_delito, set()).add(cutoff_row.fecha_corte_mindefensa)

    # Formatear respuesta con comparativas
    result_data = []
    for row in local_data:
        local_total = int(row.total)
        period_end_month = int(row.period_end_month or 12)
        yoy_total = sum(
            total for month, total in yoy_monthly.get(row.tipo_delito, {}).items()
            if month <= period_end_month
        )
        
        # Variación YoY
        yoy_var = local_total - yoy_total
        yoy_pct = year_over_year(local_total, yoy_total)
        national_benchmark = comparable_national_rate(
            year=anio,
            local_code=local_code,
            local_total=local_total,
            national_total=national_totals_by_type.get(row.tipo_delito, 0),
            covered_codes=coverage_codes_by_type.get(row.tipo_delito, set()),
            cutoffs=cutoffs_by_type.get(row.tipo_delito, set()),
        )
        territorial_benchmark = comparable_reference_rate(
            year=anio,
            local_code=local_code,
            local_total=local_total,
            reference_total=territorial_totals_by_type.get(row.tipo_delito, 0),
            expected_codes=territorial_peer_codes,
            covered_codes=territorial_codes_by_type.get(row.tipo_delito, set()),
            cutoffs=cutoffs_by_type.get(row.tipo_delito, set()),
        )
        territorial_comparison = named_territorial_comparison(
            year=anio,
            target_code=local_code,
            target_total=local_total,
            expected_codes=territorial_peer_codes,
            totals_by_code=territorial_totals_by_code_by_type.get(row.tipo_delito, {}),
            covered_codes=territorial_codes_by_type.get(row.tipo_delito, set()),
            cutoffs=cutoffs_by_type.get(row.tipo_delito, set()),
        )

        result_data.append({
            "delito": row.tipo_delito,
            "local": local_total,
            "yoy_total": yoy_total,
            "yoy_var": yoy_var,
            "yoy_pct": yoy_pct,
            "period_end_month": period_end_month,
            "rate_per_100k": national_benchmark["local_rate_per_100k"],
            "territorial_benchmark": territorial_benchmark,
            "territorial_comparison": territorial_comparison,
            "national_benchmark": national_benchmark,
        })

    # 5. Datos específicos y trazabilidad de fuente
    source_rows = db.query(
        NationalCrimeStats.source_id,
        func.max(NationalCrimeStats.fecha_corte_mindefensa).label("cutoff"),
        func.max(NationalCrimeStats.fecha_hecho).label("latest_period"),
    ).filter(
        mindefensa_source,
        NationalCrimeStats.anio == anio,
    ).group_by(NationalCrimeStats.source_id).all()
    source_ids = [row.source_id for row in source_rows]
    source_cutoffs = [row.cutoff or row.latest_period for row in source_rows if row.cutoff or row.latest_period]
    national_context = national_benchmark_guard(
        source_ids,
        max(source_cutoffs) if source_cutoffs else None,
        len(all_source_codes),
        year=anio,
        population_code=local_code,
    )
    available_benchmarks = [item["national_benchmark"] for item in result_data if item["national_benchmark"]["available"]]
    available_territorial_benchmarks = [
        item["territorial_benchmark"]
        for item in result_data
        if item["territorial_benchmark"]["available"]
    ]
    national_context["coverage"] = {
        "conductas_evaluated": len(result_data),
        "conductas_with_complete_coverage": len(available_benchmarks),
        "required_municipalities": national_context["population"]["national_universe"],
        "observed_municipalities": len(all_source_codes),
    }
    if available_benchmarks:
        national_context.update({
            "available": True,
            "status": "COMPARABLE_RATES_AVAILABLE",
            "title": "Referencia nacional verificable",
            "reason": "La tasa nacional se muestra solo en las conductas con cobertura municipal completa para el mismo ano y corte.",
        })
    national_context["territorial_reference"] = {
        "available": bool(available_territorial_benchmarks),
        "title": "Municipios de referencia poblacional de Valle del Cauca y Cauca",
        "reason": (
            "Incluye municipios con una poblacion entre 50% y 200% de la poblacion "
            "DANE del municipio consultado; se publica solo con cobertura y corte completos."
        ),
        "expected_municipalities": len(territorial_peer_codes),
        "conductas_evaluated": len(result_data),
        "conductas_with_complete_coverage": len(available_territorial_benchmarks),
    }
    national_context["dataset_scope"] = "NATIONAL_REFERENCE" if (has_compact_reference or has_reference_data) else "JAMUNDI_HISTORICAL_FALLBACK"

    fp_data = db.query(
        NationalCrimeStats.accion,
        NationalCrimeStats.institucion,
        func.sum(NationalCrimeStats.cantidad).label("total")
    ).filter(
        mindefensa_source,
        NationalCrimeStats.municipio_normalizado == target_municipio,
        NationalCrimeStats.anio == anio,
        NationalCrimeStats.tipo_delito == "Afectación Fuerza Pública"
    ).group_by(NationalCrimeStats.accion, NationalCrimeStats.institucion).all()

    fuerza_publica_summary = []
    for fp in fp_data:
        fuerza_publica_summary.append({
            "accion": fp.accion,
            "institucion": fp.institucion,
            "total": int(fp.total)
        })

    return {
        "municipio": municipio,
        "anio": anio,
        "summary": result_data,
        "trend": [{"mes": row.mes, "cantidad": int(row.total)} for row in trend_data],
        "fuerza_publica": fuerza_publica_summary,
        "context": national_context,
    }

@router.get("/municipios")
async def get_available_municipios(
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    """
    Retorna la lista de municipios que tienen datos cargados en el sistema.
    """
    municipios = db.query(
        NationalCrimeStats.codigo_dane,
        NationalCrimeStats.municipio_normalizado,
        NationalCrimeStats.municipio
    ).filter(
        NationalCrimeStats.source_id.ilike("%MINDEFENSA%")
    ).distinct().all()

    # MinDefensa puede traer variantes de escritura para un mismo municipio.
    # El codigo DANE mantiene una sola opcion y el nombre visible oficial.
    processor = NationalStatsProcessor()
    options_by_code = {}
    for municipality in municipios:
        code = normalize_municipality_code(municipality.codigo_dane)
        if not code:
            continue

        official_name = municipality_name_for_code(code)
        if not official_name:
            continue

        options_by_code[code] = {
            "id": processor.normalize_text(official_name),
            "nombre": official_name,
        }

    return sorted(options_by_code.values(), key=lambda option: option["nombre"])

@router.get("/years")
async def get_available_years(
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access),
):
    """
    Retorna la lista de años únicos que tienen datos cargados en el sistema.
    """
    from sqlalchemy import func
    anios = db.query(NationalCrimeStats.anio).filter(
        NationalCrimeStats.source_id.ilike("%MINDEFENSA%")
    ).distinct().order_by(NationalCrimeStats.anio.desc()).all()
    return [a.anio for a in anios]

@router.get("/territorial-context")
async def get_territorial_context(
    request: Request,
    fuente: str = "ASPERSION", 
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access)
):
    """
    Retorna datos agregados de contexto territorial (ej: Aspersión) para el Valle del Cauca.
    """
    await log_audit(
        db,
        "TERRITORIAL_CONTEXT_VIEW",
        actor_id=str(current_user.id),
        module="INTELLIGENCE",
        target={"fuente": fuente},
        level=2,
        request=request
    )
    # 1. Total por departamento (Valle = 76)
    valle_data = db.query(
        TerritorialContext.municipio,
        func.sum(TerritorialContext.cantidad).label("total")
    ).filter(
        TerritorialContext.fuente_id == fuente,
        TerritorialContext.codigo_depto == 76
    ).group_by(TerritorialContext.municipio).order_by(func.sum(TerritorialContext.cantidad).desc()).all()

    # 2. Serie temporal regional
    trend_valle = db.query(
        TerritorialContext.anio,
        func.sum(TerritorialContext.cantidad).label("total")
    ).filter(
        TerritorialContext.fuente_id == fuente,
        TerritorialContext.codigo_depto == 76
    ).group_by(TerritorialContext.anio).order_by(TerritorialContext.anio).all()

    return {
        "fuente": fuente,
        "region": "VALLE DEL CAUCA",
        "top_municipios": [{"municipio": r.municipio, "total": float(r.total)} for r in valle_data],
        "trend": [{"anio": r.anio, "total": float(r.total)} for r in trend_valle]
    }

@router.get("/insights")
async def get_intelligence_insights(
    request: Request,
    municipio: str = "JAMUNDI", 
    anio: int = 2025, 
    db: Session = Depends(get_db),
    current_user: User = Depends(institutional_access)
):
    """
    Genera una lectura narrativa prudente de la serie histórica local de MinDefensa.
    """
    await log_audit(
        db,
        "AI_INSIGHTS_VIEW",
        actor_id=str(current_user.id),
        module="INTELLIGENCE",
        target={"municipio": municipio, "anio": anio},
        level=2,
        request=request
    )
    try:
        from sqlalchemy import func
        
        # 1. Obtener los mismos datos que /stats para dar contexto a la IA
        processor = NationalStatsProcessor()
        target_municipio = processor.normalize_text(municipio)
        reference_source = NationalCrimeStats.source_id == "MINDEFENSA_REFERENCE"
        has_reference_data = db.query(NationalCrimeStats.id).filter(
            reference_source,
            NationalCrimeStats.municipio_normalizado == target_municipio,
            NationalCrimeStats.anio == anio,
        ).first() is not None
        mindefensa_source = reference_source if has_reference_data else NationalCrimeStats.source_id.ilike("%MINDEFENSA%")
        
        local_data = db.query(
            NationalCrimeStats.tipo_delito,
            func.sum(NationalCrimeStats.cantidad).label("total")
        ).filter(
            mindefensa_source,
            NationalCrimeStats.municipio_normalizado == target_municipio,
            NationalCrimeStats.anio == anio
        ).group_by(NationalCrimeStats.tipo_delito).all()

        # Nueva Lógica Refinada
        
        # National raw-count averages are not used as a municipal benchmark.
        yoy_data = db.query(
            NationalCrimeStats.tipo_delito,
            func.sum(NationalCrimeStats.cantidad).label("total")
        ).filter(
            mindefensa_source,
            NationalCrimeStats.municipio_normalizado == target_municipio,
            NationalCrimeStats.anio == anio - 1
        ).group_by(NationalCrimeStats.tipo_delito).all()
        yoy_dict = {row.tipo_delito: int(row.total) for row in yoy_data}

        # Construir resumen local comparable para la IA.
        stats_summary = ""
        for row in local_data:
            previous = yoy_dict.get(row.tipo_delito, 0)
            variation = year_over_year(int(row.total), previous)
            variation_text = f"{variation}% frente a {anio - 1}" if variation is not None else "sin base comparable en el ano anterior"
            stats_summary += f"- {row.tipo_delito}: {row.total} casos; {variation_text}.\n"

        if not stats_summary:
            return {"insight": "No hay suficientes datos disponibles para generar un análisis estratégico en este momento."}

        contexto = f"""
        Eres un redactor tecnico del SISC Jamundi. Redacta una lectura descriptiva
        de la serie historica local de {municipio} para el ano {anio}.

        DATOS:
        {stats_summary}

        REGLAS OBLIGATORIAS:
        1. Limitate a los conteos y variaciones entregados.
        2. No compares con Colombia, no menciones un promedio nacional y no compares municipios.
        3. No atribuyas causas, no califiques la situacion como critica y no recomiendes acciones operativas.
        4. Indica cuando no existe una base comparable en el ano anterior.
        5. Usa un tono tecnico, descriptivo y prudente, en maximo 70 palabras, sin Markdown.
        """

        try:
            # Validar proveedores configurados en api.ia
            if AI_PROVIDER == "MISTRAL":
                insight_text = await call_mistral(contexto)
            else:
                insight_text = await call_gemini(contexto)
                
            return {"insight": insight_text, "provider": AI_PROVIDER}
        except Exception as e:
            logger.error(f"Error generando insights de inteligencia (API IA): {e}")
            return {"insight": "Análisis estratégico no disponible temporalmente debido a un error de conexión con el motor de IA o al alcanzar el límite de la cuota gratuita."}

    except Exception as general_err:
        logger.error(f"Error estructurando datos locales para insights: {general_err}")
        return {"insight": "Error interno al preparar los datos estratégicos. Por favor verifique la conexión a la base de datos."}

@router.get("/public/rnmc-history")
def public_rnmc_history():
    """Return only the pre-aggregated, privacy-protected RNMC public dataset."""
    import csv
    from pathlib import Path

    dataset_path = (
        Path(__file__).resolve().parents[1]
        / "data"
        / "public"
        / "rnmc_jamundi_2017_2025_agregado.csv"
    )

    if not dataset_path.exists():
        return {
            "metadata": {
                "available": False,
                "message": "La serie publica agregada de medidas correctivas aun no ha sido cargada."
            },
            "records": []
        }

    with dataset_path.open(encoding="utf-8-sig", newline="") as stream:
        records = list(csv.DictReader(stream))

    return {
        "metadata": {
            "available": True,
            "title": "Medidas correctivas y convivencia",
            "source": "Registro Nacional de Medidas Correctivas",
            "privacy_rule": "La fuente contiene solo datos agregados. Las categorias menores a 10 registros se agrupan u ocultan.",
            "updated_at": records[0]["fecha_corte"] if records else None
        },
        "records": records
    }
