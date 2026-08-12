"""Build a public, aggregate RNMC dataset without individual records."""

from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from collections import Counter
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

MONTHS = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}
YEAR_FILE = re.compile(r"^(201[7-9]|202[0-5])(?:\s+\(\d+\))?$")


def normalize(value) -> str:
    text = "" if value is None else str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(character for character in text if not unicodedata.combining(character))
    return " ".join(text.upper().split())


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text[:10], pattern).date()
        except ValueError:
            pass
    return None


def value_at(row, indexes, column_name):
    index = indexes.get(column_name)
    return row[index] if index is not None and index < len(row) else None


def category(value, fallback="Sin clasificar"):
    text = " ".join(str(value or "").split())
    return text if text else fallback


def add_public_categories(rows, counter, indicator, privacy_rule, minimum):
    by_year = {}
    for (year, name), value in counter.items():
        by_year.setdefault(year, []).append((name, value))

    for year, categories in sorted(by_year.items()):
        other_total = 0
        for name, value in sorted(categories, key=lambda item: (-item[1], item[0])):
            if value >= minimum:
                rows.append([year, "", indicator, name, "Municipal", value, privacy_rule])
            else:
                other_total += value
        if other_total:
            rows.append([
                year,
                "",
                indicator,
                "Otros u ocultados por privacidad",
                "Municipal",
                other_total,
                privacy_rule,
            ])


def read_rnmc(input_dir: Path, municipality: str):
    municipality_key = normalize(municipality)
    records = []
    sources = []

    candidates_by_year = {}
    for path in input_dir.glob("*.xlsx"):
        match = YEAR_FILE.match(path.stem)
        if not match:
            continue
        candidates_by_year.setdefault(match.group(1), []).append(path)

    for year in sorted(candidates_by_year):
        candidates = sorted(candidates_by_year[year])
        path = next((item for item in candidates if " (" in item.stem), candidates[0])

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            indexes = {normalize(name): index for index, name in enumerate(header) if name is not None}

            required = ("LUGAR", "FECHA_HECHOS", "COMPORTAMIENTO", "MEDIDA", "ESTADO_MEDIDA")
            missing = [name for name in required if name not in indexes]
            if missing:
                raise ValueError(f"{path.name}: faltan columnas {', '.join(missing)}")

            accepted = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                place = normalize(value_at(row, indexes, "LUGAR"))
                if municipality_key not in place:
                    continue

                event_date = parse_date(value_at(row, indexes, "FECHA_HECHOS"))
                if not event_date:
                    continue

                records.append({
                    "date": event_date,
                    "behavior": category(value_at(row, indexes, "COMPORTAMIENTO")),
                    "measure": category(value_at(row, indexes, "MEDIDA")),
                    "status": category(value_at(row, indexes, "ESTADO_MEDIDA")),
                    "neighborhood": category(value_at(row, indexes, "BARRIO_HECHOS"), "Sin informacion territorial"),
                })
                accepted += 1

            sources.append((path.name, accepted))
        finally:
            workbook.close()

    return records, sources


def build_rows(records, minimum):
    dates = [record["date"] for record in records]
    cutoff = max(dates).isoformat()
    privacy_rule = f"Categorias menores a {minimum} registros se agrupan u ocultan."
    rows = []

    monthly = Counter((record["date"].year, record["date"].month) for record in records)
    for (year, month), value in sorted(monthly.items()):
        rows.append([
            year,
            f"{year}-{month:02d}",
            "Ordenes de comparendo",
            "Total",
            "Municipal",
            value,
            privacy_rule,
        ])

    behavior = Counter((record["date"].year, record["behavior"]) for record in records)
    measure = Counter((record["date"].year, record["measure"]) for record in records)
    status = Counter((record["date"].year, record["status"]) for record in records)
    neighborhood = Counter((record["date"].year, record["neighborhood"]) for record in records)

    add_public_categories(rows, behavior, "Comportamiento contrario a la convivencia", privacy_rule, minimum)
    add_public_categories(rows, measure, "Medida correctiva", privacy_rule, minimum)
    add_public_categories(rows, status, "Estado de medida", privacy_rule, minimum)
    add_public_categories(rows, neighborhood, "Barrio del hecho", privacy_rule, minimum)

    return cutoff, rows


def main():
    parser = argparse.ArgumentParser(
        description="Genera una tabla RNMC publica y agregada para SISC."
    )
    parser.add_argument("--input-dir", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--municipality", default="Jamundi")
    parser.add_argument("--minimum-public-count", default=10, type=int)
    args = parser.parse_args()

    records, sources = read_rnmc(args.input_dir, args.municipality)
    if not records:
        raise SystemExit(
            "No se encontraron registros con LUGAR que contenga el municipio indicado. "
            "Revise el filtro territorial antes de publicar."
        )

    cutoff, rows = build_rows(records, args.minimum_public_count)
    args.output.parent.mkdir(parents=True, exist_ok=True)

    with args.output.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.writer(stream)
        writer.writerow([
            "conjunto",
            "fecha_corte",
            "anio",
            "periodo",
            "indicador",
            "categoria",
            "zona_general",
            "valor",
            "fuente",
            "regla_privacidad",
        ])
        for year, period, indicator, item_category, zone, value, privacy_rule in rows:
            writer.writerow([
                "RNMC medidas correctivas",
                cutoff,
                year,
                period,
                indicator,
                item_category,
                zone,
                value,
                "Registro Nacional de Medidas Correctivas",
                privacy_rule,
            ])

    print(f"Registros individuales procesados internamente: {len(records)}")
    print(f"Filas agregadas publicas generadas: {len(rows)}")
    print(f"Archivo publico: {args.output}")
    for filename, count in sources:
        print(f"  {filename}: {count} registros de {args.municipality}")


if __name__ == "__main__":
    main()